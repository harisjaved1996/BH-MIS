import re

import openpyxl
from django.db import transaction

from academics.models import Result
from academics.services.grading import compute_grade, compute_percentage

HEADER_SCAN_ROWS = 5

# Fixed column positions (1-based) per the consolidated result sheet layout.
COL_ROLL_NO = 2
COL_NAME = 3
COL_CAMPUS = 4
COL_CITY = 5
COL_BOARD = 6
COL_TOTAL = 7
COL_OBTAINED = 8
COL_REMARKS = 11


def norm(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _to_int(value):
    if value is None or value == "":
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    if f != int(f):
        return None
    return int(f)


def _find_header_row(ws):
    for row_no, row in enumerate(
        ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True), start=1
    ):
        cells = [norm(v).lower() for v in row[: COL_REMARKS + 1]]
        if len(cells) > COL_ROLL_NO - 1 and "roll" in cells[COL_ROLL_NO - 1]:
            if len(cells) > COL_TOTAL - 1 and "total" in cells[COL_TOTAL - 1]:
                return row_no
            raise ValueError(
                "Sheet layout not recognized: found a Roll No column but column G "
                "is not Total Marks. Expected columns: Sr No, Roll No, Student Name, "
                "Campus, City, Board, Total Marks, Obtained Marks, ..., Remarks."
            )
    raise ValueError(
        "Sheet layout not recognized: no header row with 'Roll No' found in the "
        "first 5 rows. Expected columns: Sr No, Roll No, Student Name, Campus, "
        "City, Board, Total Marks, Obtained Marks, ..., Remarks."
    )


def import_results(file, session, user):
    """Parse an .xlsx result sheet and insert new results for `session`.

    Returns {"inserted", "skipped_duplicates", "ungraded", "errors": [{"row", "message"}]}.
    Existing (roll_no, session, board) rows are never overridden.
    """
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    header_row = _find_header_row(ws)

    existing = set(
        Result.objects.filter(session=session).values_list("roll_no", "board")
    )
    seen_in_file = set()

    to_create = []
    errors = []
    skipped_duplicates = []
    ungraded = []

    for offset, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True)
    ):
        row_no = header_row + 1 + offset
        values = list(row[:COL_REMARKS])
        values += [None] * (COL_REMARKS - len(values))

        roll_no = norm(values[COL_ROLL_NO - 1])
        student_name = norm(values[COL_NAME - 1])
        campus = norm(values[COL_CAMPUS - 1])
        city = norm(values[COL_CITY - 1])
        board = norm(values[COL_BOARD - 1]).upper()
        remarks = norm(values[COL_REMARKS - 1])

        if not any([roll_no, student_name, campus, city, board]):
            continue  # fully blank row

        def row_error(message):
            errors.append(
                {
                    "row": row_no,
                    "roll_no": roll_no,
                    "student_name": student_name,
                    "message": message,
                }
            )

        if not roll_no or not student_name:
            row_error("Missing roll no or student name.")
            continue

        total = _to_int(values[COL_TOTAL - 1])
        obtained = _to_int(values[COL_OBTAINED - 1])
        if total is None or obtained is None:
            row_error("Total/obtained marks must be whole numbers.")
            continue
        if total <= 0:
            row_error("Total marks must be greater than zero.")
            continue
        if obtained < 0 or obtained > total:
            row_error("Obtained marks must be between 0 and total marks.")
            continue

        key = (roll_no, board)
        if key in existing or key in seen_in_file:
            skipped_duplicates.append(
                {
                    "row": row_no,
                    "roll_no": roll_no,
                    "student_name": student_name,
                    "board": board,
                    "reason": (
                        "Duplicate within this file"
                        if key in seen_in_file
                        else f"Already exists in session {session}"
                    ),
                }
            )
            continue
        seen_in_file.add(key)

        percentage = compute_percentage(obtained, total)
        grade = compute_grade(percentage)
        if not grade:
            ungraded.append(
                {
                    "row": row_no,
                    "roll_no": roll_no,
                    "student_name": student_name,
                    "percentage": str(percentage),
                }
            )

        to_create.append(
            Result(
                roll_no=roll_no,
                student_name=student_name,
                campus=campus,
                city=city,
                board=board,
                session=session,
                total_marks=total,
                obtained_marks=obtained,
                percentage=percentage,
                grade=grade,
                remarks=remarks,
                created_by=user,
                updated_by=user,
            )
        )

    wb.close()

    with transaction.atomic():
        Result.objects.bulk_create(to_create)

    return {
        "inserted": len(to_create),
        "skipped_duplicates": skipped_duplicates,  # list of records, not inserted
        "ungraded": ungraded,  # list of records inserted without a matching grade band
        "errors": errors,  # list of rejected rows with reasons
    }
